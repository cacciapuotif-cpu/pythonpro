"""Regressioni permanenti: modello, specifica, template e import restano allineati."""

from io import BytesIO
from pathlib import Path
import sys

from openpyxl import Workbook, load_workbook
import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

sys.path.insert(0, str(Path(__file__).parent.parent))

from database import Base
import models
import schemas
from services.azienda_excel import build_workbook, import_workbook, preview_workbook
from services.azienda_field_spec import COMPANY_FIELDS, SHEET_SPECS, headers, importable_company_fields


@pytest.fixture
def db_session(tmp_path):
    engine = create_engine(
        f"sqlite:///{tmp_path / 'azienda_excel.db'}",
        connect_args={"check_same_thread": False},
    )
    testing_session = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    Base.metadata.create_all(bind=engine)
    session = testing_session()
    try:
        yield session
    finally:
        session.close()
        Base.metadata.drop_all(bind=engine)
        engine.dispose()


def _bytes(workbook):
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _header_map(worksheet):
    return {cell.value: cell.column for cell in worksheet[1]}


def _set_row(worksheet, row_number, values):
    columns = _header_map(worksheet)
    for label, value in values.items():
        worksheet.cell(row_number, columns[label], value)


def _base_workbook():
    return load_workbook(BytesIO(build_workbook().getvalue()))


def test_specifica_copre_modello_e_genera_esattamente_le_intestazioni():
    model_columns = set(models.AziendaCliente.__table__.columns.keys())
    assert {field["name"] for field in COMPANY_FIELDS if field["name"] not in {"agenzia_id", "consulente_id"}} <= model_columns
    assert {field["name"] for field in COMPANY_FIELDS} <= set(schemas.AziendaClienteBase.model_fields)

    workbook = _base_workbook()
    assert {"Istruzioni", "Aziende", "Sedi", "Conti", "Fondi"} <= set(workbook.sheetnames)
    assert [cell.value for cell in workbook["Aziende"][1]] == headers(importable_company_fields())
    assert len(workbook["Aziende"].data_validations.dataValidation) >= 3
    for sheet_name, fields in SHEET_SPECS.items():
        assert [cell.value for cell in workbook[sheet_name][1]] == headers(fields)
        assert all(cell.value for cell in workbook[sheet_name][2])
        assert any(cell.value for cell in workbook[sheet_name][3])


def test_import_completo_due_aziende_tre_sedi_reimport_e_round_trip(db_session):
    workbook = _base_workbook()
    aziende = workbook["Aziende"]
    _set_row(aziende, 4, {
        "Ragione sociale": "Alfa Formazione Srl", "Partita IVA": "11111111115",
        "Codice fiscale": "11111111115", "Codice ATECO": "85.59.20",
        "Indirizzo sede legale": "Via Alfa 1", "Comune sede legale": "Napoli",
        "CAP sede legale": "80100", "Provincia sede legale": "NA",
        "Email": "info@alfa.example", "CCNL prevalente": "Commercio",
        "Numero dipendenti": 12, "Regime aiuti predefinito": "de_minimis", "Stato azienda": "Sì",
    })
    _set_row(aziende, 5, {
        "Ragione sociale": "Beta Academy Spa", "Partita IVA": "22222222220",
        "Codice fiscale": "22222222220", "Comune sede legale": "Roma",
        "CAP sede legale": "00100", "Provincia sede legale": "RM", "Stato azienda": "Sì",
    })
    sedi = workbook["Sedi"]
    for row_number, name, city, primary in (
        (4, "Napoli Centro", "Napoli", "Sì"),
        (5, "Caserta Aule", "Caserta", "No"),
        (6, "Salerno Lab", "Salerno", "No"),
    ):
        _set_row(sedi, row_number, {
            "Partita IVA azienda": "11111111115", "Denominazione sede": name,
            "Tipo sede": "operativa", "Indirizzo": f"Via {name} 1", "Comune": city,
            "Provincia": "NA" if city in {"Napoli", "Caserta"} else "SA", "CAP": "80100",
            "Sede principale": primary,
        })
    _set_row(workbook["Conti"], 4, {
        "Partita IVA azienda": "11111111115", "Banca": "Banca Alfa",
        "IBAN": "IT60X0542811101000000123456", "Intestatario": "Alfa Formazione Srl",
        "Conto predefinito": "Sì", "Conto attivo": "Sì",
    })
    _set_row(workbook["Fondi"], 4, {
        "Partita IVA azienda": "11111111115", "Fondo interprofessionale": "FONDIMPRESA",
        "Data inizio adesione": "2024-01-01",
    })
    content = _bytes(workbook)

    preview = preview_workbook(content, db_session)
    assert preview["summary"] == {"create": 2, "update": 0, "reject": 0, "valid": 2}

    first = import_workbook(content, db_session)
    assert (first["created"], first["updated"], first["rejected"]) == (2, 0, 0)
    assert db_session.query(models.AziendaCliente).count() == 2
    assert db_session.query(models.AziendaClienteSedeOperativa).count() == 3
    assert db_session.query(models.AziendaClienteBankAccount).count() == 1

    second = import_workbook(content, db_session)
    assert (second["created"], second["updated"], second["rejected"]) == (0, 2, 0)
    assert db_session.query(models.AziendaCliente).count() == 2
    assert db_session.query(models.AziendaClienteSedeOperativa).count() == 3

    companies = db_session.query(models.AziendaCliente).all()
    exported = build_workbook(companies, reveal_sensitive=True).getvalue()
    round_trip = preview_workbook(exported, db_session)
    assert round_trip["summary"] == {"create": 0, "update": 2, "reject": 0, "valid": 2}


def test_file_misto_scarto_riga_senza_ragione_sociale_senza_placeholder(db_session):
    workbook = _base_workbook()
    _set_row(workbook["Aziende"], 4, {
        "Ragione sociale": "Gamma Srl", "Partita IVA": "33333333335", "Stato azienda": "Sì",
    })
    _set_row(workbook["Aziende"], 5, {
        "Ragione sociale": "", "Partita IVA": "44444444440", "Stato azienda": "Sì",
    })
    content = _bytes(workbook)
    preview = preview_workbook(content, db_session)
    assert preview["summary"]["valid"] == 1
    assert preview["summary"]["reject"] == 1
    assert any(error["row"] == 5 and "obbligatorio" in error["message"] for error in preview["errors"])
    result = import_workbook(content, db_session)
    assert result["created"] == 1
    assert db_session.query(models.AziendaCliente).filter(models.AziendaCliente.partita_iva == "44444444440").count() == 0


def test_intestazioni_non_riconosciute_segnala_errore_esplicito(db_session):
    """CONFUTATORE: file reale dell'utente con header propri, non dal template.

    Prima del fix questo caso tornava silenziosamente valid=0 senza errori ne'
    warning: l'utente vedeva "0 righe da importare" senza sapere perche'.
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Foglio1"
    worksheet.append(["Nome Azienda", "P.IVA", "Comune", "Osservazioni"])
    worksheet.append(["Prova Srl", "01234567890", "Napoli", "test"])

    preview = preview_workbook(_bytes(workbook), db_session)
    assert preview["summary"] == {"create": 0, "update": 0, "reject": 1, "valid": 0}
    assert len(preview["errors"]) == 1
    error = preview["errors"][0]
    assert error["sheet"] == "Foglio1"
    assert "Nome Azienda" in error["message"]
    assert "template" in error["message"].casefold()


def _sostituisci_interamente_riga(worksheet, row_number, values_by_label):
    """A differenza di _set_row, svuota prima ogni colonna: simula un utente
    che ha selezionato la riga esempio e l'ha riscritta da capo per intero,
    non solo alcune celle (che lascerebbero residui placeholder nelle altre)."""
    columns = _header_map(worksheet)
    for label, column in columns.items():
        worksheet.cell(row_number, column, values_by_label.get(label, ""))


def test_dato_reale_scritto_sopra_riga_esempio_viene_importato(db_session):
    """CONFUTATORE: caso reale utente. La riga 3 (gialla, esempio) e' la prima
    riga apparentemente libera dopo header+istruzioni: chi non nota il colore
    ci scrive sopra la propria azienda invece di partire dalla riga 4.

    Prima del fix la riga 3 veniva scartata SEMPRE per posizione, a
    prescindere dal contenuto: "Importa 0 righe valide" senza errori.
    """
    workbook = _base_workbook()
    _sostituisci_interamente_riga(workbook["Aziende"], 3, {
        "Ragione sociale": "Reale Import Srl", "Partita IVA": "00743110157",
        "Comune sede legale": "Napoli", "Stato azienda": "Sì",
    })
    preview = preview_workbook(_bytes(workbook), db_session)
    assert preview["errors"] == []
    assert preview["summary"] == {"create": 1, "update": 0, "reject": 0, "valid": 1}

    result = import_workbook(_bytes(workbook), db_session)
    assert (result["created"], result["updated"], result["rejected"]) == (1, 0, 0)
    assert db_session.query(models.AziendaCliente).filter(
        models.AziendaCliente.partita_iva == "00743110157"
    ).count() == 1


def test_riga_esempio_intatta_resta_esclusa(db_session):
    """Regressione: l'esempio non toccato in riga 3 continua a non importarsi."""
    workbook = _base_workbook()
    _set_row(workbook["Aziende"], 4, {
        "Ragione sociale": "Vera Azienda Srl", "Partita IVA": "00743110157", "Stato azienda": "Sì",
    })
    preview = preview_workbook(_bytes(workbook), db_session)
    assert preview["summary"] == {"create": 1, "update": 0, "reject": 0, "valid": 1}


def test_sede_reale_scritta_sopra_riga_esempio_viene_importata(db_session):
    """Stesso bug, verificato anche sul foglio Sedi (non solo Aziende)."""
    workbook = _base_workbook()
    _set_row(workbook["Aziende"], 4, {
        "Ragione sociale": "Vera Azienda Srl", "Partita IVA": "00743110157", "Stato azienda": "Sì",
    })
    _sostituisci_interamente_riga(workbook["Sedi"], 3, {
        "Partita IVA azienda": "00743110157", "Denominazione sede": "Sede reale utente",
        "Tipo sede": "operativa", "Comune": "Caserta",
    })
    preview = preview_workbook(_bytes(workbook), db_session)
    assert preview["errors"] == []
    result = import_workbook(_bytes(workbook), db_session)
    assert result["created"] == 1
    sedi = db_session.query(models.AziendaClienteSedeOperativa).all()
    assert any(sede.nome == "Sede reale utente" for sede in sedi)


def test_sede_principale_vuota_non_rompe_la_validazione(db_session):
    """CONFUTATORE: 'Sede principale' lasciata vuota (non e' obbligatoria) non
    deve far fallire l'import con un errore che punta al campo sbagliato.
    is_principale e' bool non Optional nello schema: None esplicito rompeva
    Pydantic con 'Input should be a valid boolean' riferito a 'sedi_operative'.
    """
    workbook = _base_workbook()
    _set_row(workbook["Aziende"], 4, {
        "Ragione sociale": "Vera Azienda Srl", "Partita IVA": "00743110157", "Stato azienda": "Sì",
    })
    _set_row(workbook["Sedi"], 4, {
        "Partita IVA azienda": "00743110157", "Denominazione sede": "Sede senza flag principale",
        "Tipo sede": "operativa",
    })
    preview = preview_workbook(_bytes(workbook), db_session)
    assert preview["errors"] == []
    assert preview["summary"]["valid"] == 1

    result = import_workbook(_bytes(workbook), db_session)
    assert result["created"] == 1
    sede = db_session.query(models.AziendaClienteSedeOperativa).filter(
        models.AziendaClienteSedeOperativa.nome == "Sede senza flag principale"
    ).first()
    assert sede is not None
    assert sede.is_principale is False


def test_vecchio_formato_resta_accettato_con_avviso_deprecazione(db_session):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Aziende"
    legacy_headers = [
        "Ragione Sociale", "Partita IVA", "Codice Fiscale", "Settore ATECO", "Attività Erogate",
        "Indirizzo Sede Legale", "Città", "CAP", "Provincia", "Email", "PEC", "Telefono",
        "Sito Web", "Sedi Operative", "Note",
    ]
    worksheet.append(legacy_headers)
    worksheet.append([
        "Legacy Srl", "11111111115", "11111111115", "85.59", "Formazione", "Via Roma 1",
        "Napoli", "80100", "NA", "legacy@example.it", "", "0810000000", "https://example.it",
        "Napoli|Via Toledo 10|Napoli|80134|NA|Aule", "Import storico",
    ])
    preview = preview_workbook(_bytes(workbook), db_session)
    assert preview["format"] == "legacy"
    assert preview["summary"]["valid"] == 1
    assert any("deprecato" in warning for warning in preview["warnings"])
