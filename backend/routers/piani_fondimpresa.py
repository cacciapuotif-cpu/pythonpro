"""Router per il modulo Piano Finanziario Fondimpresa."""

from io import BytesIO
import logging
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

import crud
import schemas
from database import get_db
from piano_fondimpresa_config import SEZIONE_TITLES

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/v1/piani-fondimpresa", tags=["Piani Fondimpresa"])


def _build_detail_response(piano) -> schemas.PianoFondimpresaDettaglio:
    riepilogo = crud.build_piano_fondimpresa_riepilogo(piano)
    totale_riferimento = riepilogo["totale_escluso_cofinanziamento"]

    dettaglio_budget = None
    if piano.dettaglio_budget:
        dettaglio_budget = schemas.DettaglioBudgetFondimpresa(
            consulenti=[
                schemas.BudgetConsulenteFondimpresa(
                    id=item.id,
                    nominativo=item.nominativo,
                    ore=float(item.ore or 0.0),
                    costo_orario=float(item.costo_orario or 0.0),
                )
                for item in piano.dettaglio_budget.consulenti
            ],
            costi_fissi=[
                schemas.BudgetCostoFissoFondimpresa(
                    id=item.id,
                    tipologia=item.tipologia,
                    parametro=item.parametro,
                    totale=float(item.totale or 0.0),
                )
                for item in piano.dettaglio_budget.costi_fissi
            ],
            margini=[
                schemas.BudgetMargineFondimpresa(
                    id=item.id,
                    tipologia=item.tipologia,
                    percentuale=float(item.percentuale or 0.0),
                    totale_riferimento=totale_riferimento,
                )
                for item in piano.dettaglio_budget.margini
            ],
        )

    return schemas.PianoFondimpresaDettaglio(
        id=piano.id,
        progetto_id=piano.progetto_id,
        anno=piano.anno,
        ente_erogatore=piano.ente_erogatore,
        tipo_conto=piano.tipo_conto,
        totale_preventivo=float(piano.totale_preventivo or 0.0),
        created_at=piano.created_at,
        updated_at=piano.updated_at,
        progetto=schemas.Project.model_validate(piano.progetto),
        voci=[
            schemas.VoceFondimpresa(
                id=voce.id,
                sezione=voce.sezione,
                voce_codice=voce.voce_codice,
                descrizione=voce.descrizione,
                note_temporali=voce.note_temporali,
                totale_voce=float(voce.totale_voce or 0.0),
                righe_nominativo=[
                    schemas.RigaNominativoFondimpresa(
                        id=row.id,
                        nominativo=row.nominativo,
                        ore=float(row.ore or 0.0),
                        costo_orario=float(row.costo_orario or 0.0),
                    )
                    for row in voce.righe_nominativo
                ],
                documenti=[
                    schemas.DocumentoFondimpresa(
                        id=documento.id,
                        tipo_documento=documento.tipo_documento,
                        numero_documento=documento.numero_documento,
                        data_documento=documento.data_documento,
                        importo_totale=float(documento.importo_totale or 0.0),
                        importo_imputato=float(documento.importo_imputato or 0.0),
                        data_pagamento=documento.data_pagamento,
                    )
                    for documento in voce.documenti
                ],
            )
            for voce in sorted(piano.voci, key=lambda item: (item.sezione, item.voce_codice))
        ],
        dettaglio_budget=dettaglio_budget,
    )


def _build_excel_workbook(piano):
    riepilogo = crud.build_piano_fondimpresa_riepilogo(piano)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Fondimpresa"

    bold = Font(bold=True)
    section_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    total_fill = PatternFill(fill_type="solid", fgColor="E5E7EB")

    headers = [
        "VOCE", "DESCRIZIONE", "TOTALE", "%", "NOMINATIVO", "ORE",
        "COSTO ORARIO", "TOT", "TIPO DOC", "NUMERO", "DATA", "IMPORTO TOTALE",
        "IMPORTO IMPUTATO", "DATA PAGAMENTO",
    ]
    for index, label in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column=index, value=label)
        cell.font = bold
        cell.fill = section_fill

    row_cursor = 4
    section_total_rows = {}
    for sezione in ["A", "B", "C", "D"]:
        sheet.cell(row=row_cursor, column=1, value=SEZIONE_TITLES[sezione]).font = bold
        sheet.cell(row=row_cursor, column=1).fill = section_fill
        section_start = row_cursor + 1
        row_cursor += 1
        for voce in sorted([item for item in piano.voci if item.sezione == sezione], key=lambda item: item.voce_codice):
            righe = voce.righe_nominativo or [None]
            docs = voce.documenti or [None]
            max_rows = max(len(righe), len(docs))
            for idx in range(max_rows):
                row = righe[idx] if idx < len(righe) else None
                doc = docs[idx] if idx < len(docs) else None
                sheet.cell(row=row_cursor, column=1, value=voce.voce_codice if idx == 0 else "")
                sheet.cell(row=row_cursor, column=2, value=voce.descrizione if idx == 0 else "")
                sheet.cell(row=row_cursor, column=3, value=float(voce.totale_voce or 0.0) if idx == 0 else "")
                if idx == 0 and riepilogo["totale_escluso_cofinanziamento"] and sezione != "B":
                    sheet.cell(row=row_cursor, column=4, value=float(voce.totale_voce or 0.0) / riepilogo["totale_escluso_cofinanziamento"])
                if row:
                    sheet.cell(row=row_cursor, column=5, value=row.nominativo)
                    sheet.cell(row=row_cursor, column=6, value=float(row.ore or 0.0))
                    sheet.cell(row=row_cursor, column=7, value=float(row.costo_orario or 0.0))
                    sheet.cell(row=row_cursor, column=8, value=float(row.totale or 0.0))
                if doc:
                    sheet.cell(row=row_cursor, column=9, value=doc.tipo_documento)
                    sheet.cell(row=row_cursor, column=10, value=doc.numero_documento)
                    sheet.cell(row=row_cursor, column=11, value=doc.data_documento)
                    sheet.cell(row=row_cursor, column=12, value=float(doc.importo_totale or 0.0))
                    sheet.cell(row=row_cursor, column=13, value=float(doc.importo_imputato or 0.0))
                    sheet.cell(row=row_cursor, column=14, value=doc.data_pagamento)
                row_cursor += 1
        sheet.cell(row=row_cursor, column=1, value=f"TOTALE {sezione}").font = bold
        sheet.cell(row=row_cursor, column=3, value=f"=SUM(C{section_start}:C{row_cursor - 1})")
        sheet.cell(row=row_cursor, column=4, value=f"=IF($C$63=0,0,C{row_cursor}/$C$63)")
        for column in range(1, 15):
            sheet.cell(row=row_cursor, column=column).fill = total_fill
            sheet.cell(row=row_cursor, column=column).font = bold
        section_total_rows[sezione] = row_cursor
        row_cursor += 1

    sheet.cell(row=63, column=1, value="TOTALE ESCLUSO COFINANZIAMENTO").font = bold
    sheet.cell(row=63, column=3, value=f"=C{section_total_rows['A']}+C{section_total_rows['C']}+C{section_total_rows['D']}")
    sheet.cell(row=65, column=1, value="TOTALE PREVENTIVO").font = bold
    sheet.cell(row=65, column=3, value=float(piano.totale_preventivo or 0.0))
    sheet.cell(row=67, column=1, value="DIFFERENZA").font = bold
    sheet.cell(row=67, column=3, value="=C65-C63")

    sheet.cell(row=67, column=8, value="DETTAGLIO BUDGET").font = bold
    sheet.cell(row=68, column=8, value="CONSULENTI").font = bold
    budget_row = 69
    if piano.dettaglio_budget:
        for item in piano.dettaglio_budget.consulenti:
            sheet.cell(row=budget_row, column=8, value=item.nominativo)
            sheet.cell(row=budget_row, column=9, value=float(item.ore or 0.0))
            sheet.cell(row=budget_row, column=10, value=float(item.costo_orario or 0.0))
            sheet.cell(row=budget_row, column=11, value=float(item.totale or 0.0))
            budget_row += 1

        budget_row += 1
        sheet.cell(row=budget_row, column=8, value="COSTI FISSI").font = bold
        budget_row += 1
        for item in piano.dettaglio_budget.costi_fissi:
            sheet.cell(row=budget_row, column=8, value=item.tipologia)
            sheet.cell(row=budget_row, column=9, value=item.parametro)
            sheet.cell(row=budget_row, column=11, value=float(item.totale or 0.0))
            budget_row += 1

        budget_row += 1
        sheet.cell(row=budget_row, column=8, value="MARGINE").font = bold
        budget_row += 1
        for item in piano.dettaglio_budget.margini:
            sheet.cell(row=budget_row, column=8, value=item.tipologia)
            sheet.cell(row=budget_row, column=9, value=float(item.percentuale or 0.0) / 100)
            sheet.cell(row=budget_row, column=11, value=float(item.totale or 0.0))
            budget_row += 1

    for column_letter in ["C", "G", "H", "L", "M", "K"]:
        for row in range(4, max(row_cursor, 87) + 1):
            sheet[f"{column_letter}{row}"].number_format = "€ #,##0.00"
    for row in range(4, max(row_cursor, 87) + 1):
        sheet[f"D{row}"].number_format = "0.00%"
        sheet[f"I{row}"].number_format = "0.00%"

    return workbook


@router.get("/", response_model=List[schemas.PianoFondimpresa])
def list_piani_fondimpresa(
    progetto_id: Optional[int] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=200),
    db: Session = Depends(get_db),
):
    return crud.get_piani_fondimpresa(db, progetto_id=progetto_id, skip=skip, limit=limit)


@router.post("/", response_model=schemas.PianoFondimpresaDettaglio, status_code=status.HTTP_201_CREATED)
def create_piano_fondimpresa(payload: schemas.PianoFondimpresaCreate, db: Session = Depends(get_db)):
    try:
        piano = crud.create_piano_fondimpresa(db, payload)
        return _build_detail_response(piano)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@router.get("/{piano_id}", response_model=schemas.PianoFondimpresaDettaglio)
def get_piano_fondimpresa(piano_id: int, db: Session = Depends(get_db)):
    piano = crud.get_piano_fondimpresa(db, piano_id)
    if not piano:
        raise HTTPException(status_code=404, detail="Piano Fondimpresa non trovato")
    return _build_detail_response(piano)


@router.put("/{piano_id}/voci", response_model=schemas.PianoFondimpresaDettaglio)
def update_voci_fondimpresa(piano_id: int, payload: schemas.PianoFondimpresaBulkUpdate, db: Session = Depends(get_db)):
    try:
        piano = crud.bulk_upsert_voci_fondimpresa(db, piano_id, payload)
        if not piano:
            raise HTTPException(status_code=404, detail="Piano Fondimpresa non trovato")
        return _build_detail_response(piano)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@router.put("/{piano_id}/documenti", response_model=schemas.PianoFondimpresaDettaglio)
def update_documenti_fondimpresa(piano_id: int, payload: schemas.PianoFondimpresaDocumentiBulkUpdate, db: Session = Depends(get_db)):
    try:
        piano = crud.bulk_upsert_documenti_fondimpresa(db, piano_id, payload)
        if not piano:
            raise HTTPException(status_code=404, detail="Piano Fondimpresa non trovato")
        return _build_detail_response(piano)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@router.put("/{piano_id}/dettaglio-budget", response_model=schemas.PianoFondimpresaDettaglio)
def put_dettaglio_budget_fondimpresa(piano_id: int, payload: schemas.DettaglioBudgetFondimpresaUpdate, db: Session = Depends(get_db)):
    piano = crud.update_dettaglio_budget_fondimpresa(db, piano_id, payload)
    if not piano:
        raise HTTPException(status_code=404, detail="Piano Fondimpresa non trovato")
    return _build_detail_response(piano)


@router.get("/{piano_id}/riepilogo", response_model=schemas.PianoFondimpresaRiepilogo)
def get_riepilogo_fondimpresa(piano_id: int, db: Session = Depends(get_db)):
    piano = crud.get_piano_fondimpresa(db, piano_id)
    if not piano:
        raise HTTPException(status_code=404, detail="Piano Fondimpresa non trovato")
    return crud.build_piano_fondimpresa_riepilogo(piano)


@router.get("/{piano_id}/export-excel")
def export_piano_fondimpresa_excel(piano_id: int, db: Session = Depends(get_db)):
    piano = crud.get_piano_fondimpresa(db, piano_id)
    if not piano:
        raise HTTPException(status_code=404, detail="Piano Fondimpresa non trovato")

    workbook = _build_excel_workbook(piano)
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    filename = f"piano_fondimpresa_{piano.progetto.name.replace(' ', '_')}_{piano.anno}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
    )
