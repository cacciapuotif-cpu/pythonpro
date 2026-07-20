"""
Router per gestione piani finanziari e relative voci.
Pattern allineato agli altri router CRUD del progetto.
"""

from io import BytesIO
from typing import List, Optional
import logging

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

import crud
import models
import schemas
from auth import get_current_user
from database import get_db
from services.massimali import FONTE_REGOLA_AVVISO, get_massimale_effettivo
from services.piano_templates import crea_piano_da_template

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/v1/piani-finanziari", tags=["Piani Finanziari"])


def _validate_massimale_voce(db: Session, piano, voce_payload) -> None:
    """E1.3/E1.4: delega a services.massimali (precedenza regola avviso
    validata > MassimaleFondo); mantiene il 422 e cita fonte e articolo."""
    tariffa = getattr(voce_payload, "tariffa_oraria", None)
    if tariffa is None:
        return
    categoria = (getattr(voce_payload, "categoria", None) or "").lower()
    if categoria not in {"docenza", "tutoraggio"}:
        return
    effettivo = get_massimale_effettivo(db, piano, categoria)
    if effettivo is None:
        return
    if float(tariffa) > float(effettivo.limite):
        if effettivo.fonte == FONTE_REGOLA_AVVISO:
            detail = (
                f"Costo orario {categoria} ({float(tariffa):.2f}) supera il massimale "
                f"da regola avviso validata ({float(effettivo.limite):.2f})"
            )
            if effettivo.riferimento_articolo:
                detail += f" — rif. {effettivo.riferimento_articolo}"
        else:
            detail = (
                f"Costo orario {categoria} ({float(tariffa):.2f}) supera massimale "
                f"fondo ({float(effettivo.limite):.2f})"
            )
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=detail,
        )


def _build_excel_workbook(piano, riepilogo: dict) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Piano Finanziario"

    bold = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    total_fill = PatternFill(fill_type="solid", fgColor="E5E7EB")

    sheet["A1"] = "PIANO FINANZIARIO"
    sheet["A1"].font = bold
    sheet["A2"] = "Progetto"
    sheet["B2"] = piano.progetto.name if getattr(piano, "progetto", None) else f"Progetto {piano.progetto_id}"
    sheet["A3"] = "Ente erogatore"
    sheet["B3"] = piano.ente_erogatore or ""
    sheet["A4"] = "Avviso"
    sheet["B4"] = piano.avviso or ""
    sheet["A5"] = "Anno"
    sheet["B5"] = piano.anno

    headers = [
        "MACROVOCE",
        "CODICE",
        "DESCRIZIONE",
        "PROGETTO",
        "EDIZIONE",
        "ORE",
        "IMPORTO PREVENTIVO",
        "IMPORTO CONSUNTIVO",
        "IMPORTO PRESENTATO",
    ]
    for index, label in enumerate(headers, start=1):
        cell = sheet.cell(row=8, column=index, value=label)
        cell.font = bold
        cell.fill = header_fill

    rows = crud.build_effective_piano_rows(piano, db=None)
    row_cursor = 9
    for row in sorted(rows, key=lambda item: (item["macrovoce"], item["voce_codice"], item.get("progetto_label") or "", item.get("edizione_label") or "")):
        sheet.cell(row=row_cursor, column=1, value=row["macrovoce"])
        sheet.cell(row=row_cursor, column=2, value=row["voce_codice"])
        sheet.cell(row=row_cursor, column=3, value=row["descrizione"])
        sheet.cell(row=row_cursor, column=4, value=row.get("progetto_label") or "")
        sheet.cell(row=row_cursor, column=5, value=row.get("edizione_label") or "")
        sheet.cell(row=row_cursor, column=6, value=float(row.get("ore") or 0.0))
        sheet.cell(row=row_cursor, column=7, value=float(row.get("importo_preventivo") or 0.0))
        sheet.cell(row=row_cursor, column=8, value=float(row.get("importo_consuntivo") or 0.0))
        sheet.cell(row=row_cursor, column=9, value=float(row.get("importo_presentato") or 0.0))
        row_cursor += 1

    row_cursor += 1
    summary_headers = ["MACROVOCE", "TITOLO", "PREVENTIVO", "CONSUNTIVO", "% PREVENTIVO", "% CONSUNTIVO", "ALERT"]
    for index, label in enumerate(summary_headers, start=1):
        cell = sheet.cell(row=row_cursor, column=index, value=label)
        cell.font = bold
        cell.fill = total_fill
    row_cursor += 1

    for item in riepilogo.get("macrovoci", []):
        sheet.cell(row=row_cursor, column=1, value=item["macrovoce"])
        sheet.cell(row=row_cursor, column=2, value=item["titolo"])
        sheet.cell(row=row_cursor, column=3, value=float(item.get("importo_preventivo") or 0.0))
        sheet.cell(row=row_cursor, column=4, value=float(item.get("importo_consuntivo") or 0.0))
        sheet.cell(row=row_cursor, column=5, value=float(item.get("percentuale_preventivo") or 0.0) / 100)
        sheet.cell(row=row_cursor, column=6, value=float(item.get("percentuale_consuntivo") or 0.0) / 100)
        sheet.cell(row=row_cursor, column=7, value=item.get("alert_level") or "ok")
        row_cursor += 1

    row_cursor += 1
    totals = [
        ("Totale preventivo", float(riepilogo.get("totale_preventivo") or 0.0)),
        ("Totale consuntivo", float(riepilogo.get("totale_consuntivo") or 0.0)),
        ("Contributo richiesto", float(riepilogo.get("contributo_richiesto") or 0.0)),
        ("Cofinanziamento", float(riepilogo.get("cofinanziamento") or 0.0)),
    ]
    for label, value in totals:
        sheet.cell(row=row_cursor, column=1, value=label).font = bold
        sheet.cell(row=row_cursor, column=2, value=value)
        row_cursor += 1

    for column_letter in ["G", "H", "I", "B", "C", "D"]:
        for row in range(9, row_cursor + 2):
            sheet[f"{column_letter}{row}"].number_format = "EUR #,##0.00"
    for column_letter in ["E", "F"]:
        for row in range(9, row_cursor + 2):
            sheet[f"{column_letter}{row}"].number_format = "0.00%"

    return workbook


@router.get("/", response_model=List[schemas.PianoFinanziarioWithVoci])
def get_piani_finanziari(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=200),
    progetto_id: Optional[int] = Query(None),
    stato: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    return crud.get_piani_finanziari(
        db,
        skip=skip,
        limit=limit,
        progetto_id=progetto_id,
        stato=stato,
    )


@router.get("/violazioni-massimali")
def report_violazioni_massimali(db: Session = Depends(get_db)):
    """DOM-10 (GATE W1.4): report delle violazioni massimali ESISTENTI nei dati.

    - violazioni: tariffe (assignment/voce) oltre il massimale configurato
      per (tipo_fondo, anno) del piano;
    - non_verificabili: tariffe docenza/tutoraggio che nessun massimale può
      verificare (fondo senza riga massimali per l'anno) e voci con importi
      ma categoria assente (esenti dal check).
    """
    violazioni = []
    non_verificabili = []

    piani = db.query(models.PianoFinanziario).all()
    for piano in piani:
        assignments = db.query(models.Assignment).filter(
            models.Assignment.project_id == piano.progetto_id,
            models.Assignment.is_active == True,
        ).all()
        for a in assignments:
            categoria = crud._derive_categoria_from_role(a.role)
            if categoria not in ("docenza", "tutoraggio"):
                continue
            limite = crud.get_massimale_orario(db, piano.tipo_fondo, piano.anno, categoria)
            entry = {
                "piano_id": piano.id,
                "entita": "assignment",
                "entita_id": a.id,
                "categoria": categoria,
                "tariffa": float(a.hourly_rate or 0),
            }
            if limite is None:
                entry["motivo"] = (
                    f"nessun massimale configurato per fondo '{piano.tipo_fondo}' anno {piano.anno}"
                )
                non_verificabili.append(entry)
            elif float(a.hourly_rate or 0) > float(limite):
                entry["massimale"] = float(limite)
                violazioni.append(entry)

        for voce in piano.voci:
            tariffa = float(voce.tariffa_oraria or 0)
            if tariffa <= 0:
                continue
            categoria = (voce.categoria or "").lower()
            entry = {
                "piano_id": piano.id,
                "entita": "voce",
                "entita_id": voce.id,
                "categoria": categoria or None,
                "tariffa": tariffa,
            }
            if categoria not in ("docenza", "tutoraggio"):
                if not categoria:
                    entry["motivo"] = "categoria assente: voce esente dal check massimali"
                    non_verificabili.append(entry)
                continue
            limite = crud.get_massimale_orario(db, piano.tipo_fondo, piano.anno, categoria)
            if limite is None:
                entry["motivo"] = (
                    f"nessun massimale configurato per fondo '{piano.tipo_fondo}' anno {piano.anno}"
                )
                non_verificabili.append(entry)
            elif tariffa > float(limite):
                entry["massimale"] = float(limite)
                violazioni.append(entry)

    return {
        "violazioni": violazioni,
        "non_verificabili": non_verificabili,
        "totale_violazioni": len(violazioni),
        "totale_non_verificabili": len(non_verificabili),
    }


@router.get("/{piano_id}", response_model=schemas.PianoFinanziarioWithVoci)
def get_piano_finanziario(
    piano_id: int,
    db: Session = Depends(get_db),
):
    piano = crud.get_piano_finanziario(db, piano_id)
    if not piano:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Piano finanziario non trovato",
        )
    return piano


@router.post(
    "/from-template",
    response_model=schemas.PianoFinanziarioWithVoci,
    status_code=status.HTTP_201_CREATED,
)
def create_piano_finanziario_da_template(
    payload: schemas.PianoDaTemplateCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """FASE E1 (E1.2/E1.5): creazione guidata del piano da un
    ``PianoFinanziarioTemplate``. Il percorso libero ``POST /`` resta invariato.
    RBAC: stesso gate router-level di ``POST /`` (matrice operational:
    admin+operatore sui metodi non-safe, vedi ``auth.rbac_allowed_roles``)."""
    try:
        piano = crea_piano_da_template(
            db,
            template_id=payload.template_id,
            progetto_id=payload.progetto_id,
            testata=payload.model_dump(exclude={"template_id", "progetto_id"}),
            user=current_user,
        )
        logger.info(
            "Created piano finanziario %s da template %s", piano.id, payload.template_id
        )
        return piano
    except LookupError as exc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))


@router.post("/", response_model=schemas.PianoFinanziarioWithVoci, status_code=status.HTTP_201_CREATED)
def create_piano_finanziario(
    piano: schemas.PianoFinanziarioCreate,
    db: Session = Depends(get_db),
):
    try:
        db_piano = crud.create_piano_finanziario(db, piano)
        logger.info("Created piano finanziario: ID %s", db_piano.id)
        return db_piano
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(exc),
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Error creating piano finanziario: %s", exc)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Errore nella creazione del piano finanziario",
        )


@router.put("/{piano_id}", response_model=schemas.PianoFinanziarioWithVoci)
def update_piano_finanziario(
    piano_id: int,
    piano: schemas.PianoFinanziarioUpdate,
    db: Session = Depends(get_db),
):
    try:
        existing = crud.get_piano_finanziario(db, piano_id)
        if not existing:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Piano finanziario non trovato",
            )

        updated = crud.update_piano_finanziario(db, piano_id, piano)
        return updated
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(exc),
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Error updating piano finanziario %s: %s", piano_id, exc)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Errore nell'aggiornamento del piano finanziario",
        )


@router.delete("/{piano_id}")
def delete_piano_finanziario(
    piano_id: int,
    soft_delete: bool = True,
    db: Session = Depends(get_db),
):
    try:
        existing = crud.get_piano_finanziario(db, piano_id)
        if not existing:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Piano finanziario non trovato",
            )

        crud.delete_piano_finanziario(db, piano_id, soft_delete=soft_delete)
        return {
            "message": "Piano finanziario eliminato con successo" if not soft_delete else "Piano finanziario chiuso con successo",
            "piano_id": piano_id,
            "soft_delete": soft_delete,
        }
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(exc),
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Error deleting piano finanziario %s: %s", piano_id, exc)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Errore nell'eliminazione del piano finanziario",
        )


@router.get("/{piano_id}/voci", response_model=List[schemas.VocePianoFinanziario])
def get_voci_piano(
    piano_id: int,
    db: Session = Depends(get_db),
):
    piano = crud.get_piano_finanziario(db, piano_id)
    if not piano:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Piano finanziario non trovato",
        )
    return crud.get_voci_piano(db, piano_id)


@router.post("/{piano_id}/voci", response_model=schemas.VocePianoFinanziario, status_code=status.HTTP_201_CREATED)
def create_voce_piano(
    piano_id: int,
    voce: schemas.VocePianoFinanziarioCreate,
    db: Session = Depends(get_db),
):
    try:
        if voce.piano_id != piano_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Il piano_id del body non coincide con il path",
            )

        piano = crud.get_piano_finanziario(db, piano_id)
        if not piano:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Piano finanziario non trovato")
        _validate_massimale_voce(db, piano, voce)
        db_voce = crud.create_voce_piano(db, voce)
        logger.info("Created voce piano finanziario: ID %s", db_voce.id)
        return db_voce
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(exc),
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Error creating voce for piano %s: %s", piano_id, exc)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Errore nella creazione della voce piano",
        )


@router.put("/{piano_id}/voci/{voce_id}", response_model=schemas.VocePianoFinanziario)
def update_voce_piano(
    piano_id: int,
    voce_id: int,
    voce: schemas.VocePianoFinanziarioUpdate,
    db: Session = Depends(get_db),
):
    try:
        piano = crud.get_piano_finanziario(db, piano_id)
        if not piano:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Piano finanziario non trovato",
            )

        existing_voce = crud.get_voce_piano(db, voce_id)
        if not existing_voce or existing_voce.piano_id != piano_id:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Voce piano non trovata",
            )

        merged_payload = type("VocePayload", (), {})()
        for attr in ("tariffa_oraria", "categoria"):
            value = getattr(voce, attr, None)
            setattr(merged_payload, attr, value if value is not None else getattr(existing_voce, attr, None))
        _validate_massimale_voce(db, piano, merged_payload)
        updated = crud.update_voce_piano(db, voce_id, voce)
        return updated
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(exc),
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Error updating voce %s for piano %s: %s", voce_id, piano_id, exc)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Errore nell'aggiornamento della voce piano",
        )


@router.delete("/{piano_id}/voci/{voce_id}")
def delete_voce_piano(
    piano_id: int,
    voce_id: int,
    db: Session = Depends(get_db),
):
    try:
        piano = crud.get_piano_finanziario(db, piano_id)
        if not piano:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Piano finanziario non trovato",
            )

        existing_voce = crud.get_voce_piano(db, voce_id)
        if not existing_voce or existing_voce.piano_id != piano_id:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Voce piano non trovata",
            )

        crud.delete_voce_piano(db, voce_id)
        return {
            "message": "Voce piano eliminata con successo",
            "piano_id": piano_id,
            "voce_id": voce_id,
        }
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(exc),
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Error deleting voce %s for piano %s: %s", voce_id, piano_id, exc)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Errore nell'eliminazione della voce piano",
        )


@router.put("/{piano_id}/voci", response_model=schemas.PianoFinanziarioWithVoci)
def bulk_update_voci_piano(
    piano_id: int,
    payload: schemas.PianoFinanziarioBulkUpdate,
    db: Session = Depends(get_db),
):
    try:
        piano = crud.bulk_upsert_voci_piano(db, piano_id, payload)
        if not piano:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Piano finanziario non trovato",
            )
        return piano
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(exc),
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Error bulk updating voci for piano %s: %s", piano_id, exc)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Errore nell'aggiornamento massivo delle voci piano",
        )


@router.get("/{piano_id}/riepilogo", response_model=schemas.PianoFinanziarioRiepilogo)
def get_riepilogo_piano(piano_id: int, db: Session = Depends(get_db)):
    piano = crud.get_piano_finanziario(db, piano_id)
    if not piano:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Piano finanziario non trovato",
        )
    return crud.build_piano_finanziario_riepilogo(piano, db=db)


@router.get("/{piano_id}/riepilogo-budget", response_model=schemas.PianoFinanziarioRiepilogo)
def get_riepilogo_budget_piano(piano_id: int, db: Session = Depends(get_db)):
    piano = crud.get_piano_finanziario(db, piano_id)
    if not piano:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Piano finanziario non trovato",
        )
    return crud.build_piano_finanziario_riepilogo(piano, db=db)


@router.post("/assignments/{assignment_id}/collega-mansione", response_model=schemas.VocePianoFinanziario)
def collega_assignment_a_piano(assignment_id: int, db: Session = Depends(get_db)):
    try:
        voce = crud.collega_assegnazione_a_piano(db, assignment_id)
        if not voce:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Assegnazione o piano finanziario non trovato",
            )
        return voce
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))


@router.post("/voci/{voce_id}/aggiorna-da-presenze", response_model=schemas.VocePianoFinanziario)
def aggiorna_voce_piano_da_presenze(voce_id: int, db: Session = Depends(get_db)):
    voce = crud.aggiorna_voce_da_presenze(db, voce_id)
    if not voce:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Voce piano non trovata")
    return voce


@router.get("/{piano_id}/export-excel")
def export_piano_finanziario_excel(piano_id: int, db: Session = Depends(get_db)):
    piano = crud.get_piano_finanziario(db, piano_id)
    if not piano:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Piano finanziario non trovato",
        )

    riepilogo = crud.build_piano_finanziario_riepilogo(piano, db=db)
    workbook = _build_excel_workbook(piano, riepilogo)
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    project_name = (piano.progetto.name if getattr(piano, "progetto", None) else f"progetto_{piano.progetto_id}").replace(" ", "_")
    filename = f"piano_finanziario_{project_name}_{piano.anno}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
