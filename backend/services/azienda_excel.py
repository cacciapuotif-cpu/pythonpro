"""Template, preview, import ed export aziende dalla specifica canonica."""

from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from io import BytesIO
import re
from urllib.parse import urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import ValidationError
from sqlalchemy import func
from sqlalchemy.orm import Session

import crud
import models
import schemas
from services.azienda_field_spec import (
    COMPANY_FIELDS,
    GROUPS,
    PROVINCE,
    SHEET_SPECS,
    SPEC_VERSION,
    field_by_label,
    headers,
    importable_company_fields,
)


EXAMPLE_NAME = "ESEMPIO - NON IMPORTARE"
LEGACY_HEADERS = {
    "Ragione Sociale": "ragione_sociale",
    "Partita IVA": "partita_iva",
    "Codice Fiscale": "codice_fiscale",
    "Settore ATECO": "settore_ateco",
    "Attività Erogate": "attivita_erogate",
    "Indirizzo Sede Legale": "indirizzo",
    "Città": "citta",
    "CAP": "cap",
    "Provincia": "provincia",
    "Email": "email",
    "PEC": "pec",
    "Telefono": "telefono",
    "Sito Web": "sito_web",
    "Sedi Operative": "sedi_operative_raw",
    "Note": "note",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
FORMAT_FILL = PatternFill("solid", fgColor="D9EAF7")
EXAMPLE_FILL = PatternFill("solid", fgColor="FFF2CC")
ERROR_FILL = PatternFill("solid", fgColor="FCE8E6")


def _format_instruction(field):
    required = "OBBLIGATORIO" if field["required"] else "facoltativo"
    validation = field.get("validation") or field["type"]
    if field.get("options"):
        validation = "valori: " + ", ".join(field["options"][:8])
        if len(field["options"]) > 8:
            validation += "… (menu a discesa)"
    return f"{required}; {validation}"


def _example_company():
    return {
        "ragione_sociale": EXAMPLE_NAME,
        "natura_giuridica": "S.r.l.",
        "partita_iva": "12345678903",
        "codice_fiscale": "12345678903",
        "settore_ateco": "85.59.20",
        "settore_codice": "FORMAZIONE",
        "settore_descrizione": "Corsi di formazione e aggiornamento professionale",
        "attivita_erogate": "Formazione finanziata e consulenza",
        "indirizzo": "Via Roma 1",
        "citta": "Napoli",
        "cap": "80100",
        "provincia": "NA",
        "email": "info@esempio.invalid",
        "pec": "esempio@pec.invalid",
        "telefono": "0811234567",
        "sito_web": "https://esempio.invalid",
        "linkedin_url": "https://linkedin.com/company/esempio",
        "facebook_url": "https://facebook.com/esempio",
        "instagram_url": "https://instagram.com/esempio",
        "ccnl_prevalente": "Commercio",
        "num_dipendenti": 25,
        "matricola_inps": "1234567890",
        "anno_adesione": "2024",
        "regime_aiuto_default": "de_minimis",
        "legale_rappresentante_nome": "Mario",
        "legale_rappresentante_cognome": "Rossi",
        "legale_rappresentante_codice_fiscale": "RSSMRA80A01F839Q",
        "legale_rappresentante_email": "mario.rossi@esempio.invalid",
        "legale_rappresentante_telefono": "3330000000",
        "referente_nome": "Anna",
        "referente_cognome": "Verdi",
        "referente_ruolo": "Responsabile formazione",
        "referente_email": "anna.verdi@esempio.invalid",
        "referente_telefono": "3331111111",
        "note": "Riga dimostrativa: non viene importata",
        "attivo": "Sì",
    }


EXAMPLES = {
    "Aziende": _example_company(),
    "Sedi": {
        "partita_iva": "12345678903", "nome": "Sede Napoli Centro", "tipo": "operativa",
        "indirizzo": "Via Toledo 10", "citta": "Napoli", "provincia": "NA", "cap": "80134",
        "email": "napoli@esempio.invalid", "telefono": "0811111111", "is_principale": "Sì",
        "note": "Riga dimostrativa",
    },
    "Conti": {
        "partita_iva": "12345678903", "banca": "Banca Esempio", "agenzia": "Filiale Napoli",
        "iban": "IT60X0542811101000000123456", "bic_swift": "BPPIITRRXXX",
        "intestatario": "Azienda Esempio S.r.l.", "is_predefinito": "Sì", "is_active": "Sì",
        "note": "Riga dimostrativa",
    },
    "Fondi": {
        "partita_iva": "12345678903", "fondo": "FONDIMPRESA", "data_inizio": "2024-01-01",
        "data_fine": "", "note": "Riga dimostrativa",
    },
}


def _write_sheet(workbook, name, fields, data_rows, *, include_example):
    worksheet = workbook.create_sheet(name)
    worksheet.freeze_panes = "A3"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(fields))}1"
    worksheet.append(headers(fields))
    worksheet.append([_format_instruction(field) for field in fields])
    if include_example:
        example = EXAMPLES[name]
        worksheet.append([example.get(field["name"], "") for field in fields])

    for row in data_rows:
        worksheet.append([row.get(field["name"], "") for field in fields])

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for cell in worksheet[2]:
        cell.fill = FORMAT_FILL
        cell.font = Font(italic=True, color="375A7F")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    if include_example:
        for cell in worksheet[3]:
            cell.fill = EXAMPLE_FILL
            cell.font = Font(italic=True)

    worksheet.row_dimensions[1].height = 32
    worksheet.row_dimensions[2].height = 46
    for index, field in enumerate(fields, start=1):
        width = max(16, min(36, len(field["label"]) + 4))
        worksheet.column_dimensions[get_column_letter(index)].width = width
        options = field.get("options") or []
        if not options:
            continue
        if tuple(options) == tuple(PROVINCE):
            formula = "=Valori!$A$2:$A${}".format(len(PROVINCE) + 1)
        else:
            key = "_".join(re.sub(r"[^A-Za-z0-9]+", "_", item) for item in options)
            column = workbook["Valori"].max_column + 1
            value_sheet = workbook["Valori"]
            value_sheet.cell(1, column, key)
            for option_index, option in enumerate(options, start=2):
                value_sheet.cell(option_index, column, option)
            formula = "={}!${}${}:${}${}".format(
                quote_sheetname("Valori"), get_column_letter(column), 2,
                get_column_letter(column), len(options) + 1,
            )
        validation = DataValidation(type="list", formula1=formula, allow_blank=not field["required"])
        validation.error = "Seleziona uno dei valori previsti"
        validation.errorTitle = "Valore non valido"
        worksheet.add_data_validation(validation)
        validation.add(f"{get_column_letter(index)}3:{get_column_letter(index)}1048576")
    return worksheet


def _display_value(company, field, *, reveal_sensitive=False):
    name = field["name"]
    if name == "agenzia_id":
        return company.agenzia.nome if company.agenzia else ""
    if name == "consulente_id":
        return f"{company.consulente.cognome} {company.consulente.nome}" if company.consulente else ""
    value = getattr(company, name, None)
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "Sì" if value else "No"
    return value if value is not None else ""


def build_workbook(companies=None, *, include_example=True, reveal_sensitive=False):
    workbook = Workbook()
    workbook.remove(workbook.active)
    values = workbook.create_sheet("Valori")
    values.sheet_state = "hidden"
    values.cell(1, 1, "Province")
    for index, province in enumerate(PROVINCE, start=2):
        values.cell(index, 1, province)

    instructions = workbook.create_sheet("Istruzioni")
    instructions.append(["Template anagrafica aziende", f"Versione specifica: {SPEC_VERSION}"])
    instructions.append(["Uso", "Compila da riga 4. La riga gialla è solo un esempio e non viene importata."])
    instructions.append(["Aggiornamenti", "La Partita IVA identifica l'azienda: reimportare aggiorna, non duplica."])
    instructions.append(["Sedi / Conti / Fondi", "Una riga per elemento, collegata tramite Partita IVA azienda."])
    instructions.append(["Compatibilità", "Il vecchio foglio unico con Sedi Operative resta accettato ma è deprecato."])
    instructions.column_dimensions["A"].width = 24
    instructions.column_dimensions["B"].width = 100

    company_rows = []
    site_rows = []
    account_rows = []
    fund_rows = []
    for company in companies or []:
        company_rows.append({field["name"]: _display_value(company, field) for field in importable_company_fields()})
        for site in company.sedi_operative:
            site_rows.append({
                "partita_iva": company.partita_iva or "", "nome": site.nome, "tipo": site.tipo,
                "indirizzo": site.indirizzo or "", "citta": site.citta or "", "provincia": site.provincia or "",
                "cap": site.cap or "", "email": site.email or "", "telefono": site.telefono or "",
                "is_principale": "Sì" if site.is_principale else "No", "note": site.note or "",
            })
        for account in company.conti_correnti:
            account_rows.append({
                "partita_iva": company.partita_iva or "", "banca": account.banca or "", "agenzia": account.agenzia or "",
                "iban": account.iban if reveal_sensitive else account.iban_masked, "bic_swift": account.bic_swift or "",
                "intestatario": account.intestatario, "is_predefinito": "Sì" if account.is_predefinito else "No",
                "is_active": "Sì" if account.is_active else "No", "note": account.note or "",
            })
        for membership in company.fund_memberships:
            fund_rows.append({
                "partita_iva": company.partita_iva or "", "fondo": membership.fondo,
                "data_inizio": membership.data_inizio.date().isoformat() if isinstance(membership.data_inizio, datetime) else membership.data_inizio.isoformat(),
                "data_fine": (membership.data_fine.date().isoformat() if isinstance(membership.data_fine, datetime) else membership.data_fine.isoformat()) if membership.data_fine else "",
                "note": membership.note or "",
            })

    _write_sheet(workbook, "Aziende", importable_company_fields(), company_rows, include_example=include_example)
    rows_by_sheet = {"Sedi": site_rows, "Conti": account_rows, "Fondi": fund_rows}
    for name, fields in SHEET_SPECS.items():
        _write_sheet(workbook, name, fields, rows_by_sheet[name], include_example=include_example)
    workbook["Aziende"].conditional_formatting.add(
        "A4:A1048576", FormulaRule(formula=['A4=""'], fill=ERROR_FILL)
    )
    workbook.active = workbook.sheetnames.index("Istruzioni")
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _piva(value):
    clean = re.sub(r"\s+", "", _text(value)).removeprefix("IT").removeprefix("it")
    return clean.zfill(11) if clean.isdigit() and len(clean) < 11 else clean


def _boolean(value, default=False):
    clean = _text(value).casefold()
    if clean in {"sì", "si", "yes", "true", "1", "attivo", "attiva"}:
        return True
    if clean in {"no", "false", "0", "inattivo", "inattiva"}:
        return False
    if clean == "":
        return default
    raise ValueError("atteso Sì o No")


def _date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    return datetime.fromisoformat(_text(value)).replace(tzinfo=None)


def _validate_value(field, value):
    if value in (None, ""):
        if field["required"]:
            raise ValueError("campo obbligatorio")
        return None
    validation = field.get("validation") or ""
    text_value = _text(value)
    if validation == "partita_iva":
        return schemas._validate_piva(_piva(value))
    if validation == "codice_fiscale":
        clean = re.sub(r"\s+", "", text_value).upper()
        if not ((len(clean) == 11 and clean.isdigit()) or (len(clean) == 16 and clean.isalnum())):
            raise ValueError("Codice fiscale: 11 cifre o 16 caratteri alfanumerici")
        return clean
    if validation == "cap" and not re.fullmatch(r"\d{5}", text_value.zfill(5)):
        raise ValueError("CAP: 5 cifre")
    if validation == "cap":
        return text_value.zfill(5)
    if validation == "provincia":
        clean = text_value.upper()
        if clean not in PROVINCE:
            raise ValueError("Provincia non valida")
        return clean
    if validation == "email" and not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", text_value):
        raise ValueError("Email non valida")
    if validation == "url":
        parsed = urlparse(text_value)
        if parsed.scheme not in {"http", "https"} or not parsed.netloc:
            raise ValueError("URL non valido: usare http:// o https://")
    if validation == "iban":
        return schemas.validate_iban(text_value)
    if validation == "bic":
        clean = text_value.replace(" ", "").upper()
        if len(clean) not in {8, 11} or not clean.isalnum():
            raise ValueError("BIC/SWIFT: 8 o 11 caratteri alfanumerici")
        return clean
    if validation == "year" and not re.fullmatch(r"\d{4}", text_value):
        raise ValueError("Anno: quattro cifre")
    if validation.startswith("date:"):
        return _date(value)
    if validation == "boolean":
        return _boolean(value)
    if validation == "choice":
        if text_value not in field["options"]:
            raise ValueError("Valore ammesso: " + ", ".join(field["options"]))
        return text_value
    if validation == "integer>=0":
        parsed = int(value)
        if parsed < 0:
            raise ValueError("Il valore non può essere negativo")
        return parsed
    max_match = re.search(r"max:(\d+)", validation)
    if max_match and len(text_value) > int(max_match.group(1)):
        raise ValueError(f"Massimo {max_match.group(1)} caratteri")
    return text_value


def _resolve_foreign_key(db, field, value):
    if value in (None, ""):
        return None
    clean = _text(value).strip()
    if field["resolver"] == "agenzia_nome":
        row = db.query(models.Agenzia).filter(func.lower(models.Agenzia.nome) == clean.lower()).first()
        if not row:
            raise ValueError(f"Agenzia non trovata: {clean}")
        return row.id
    if field["resolver"] == "consulente_nome":
        rows = db.query(models.Consulente).all()
        matches = [row for row in rows if clean.casefold() in {
            f"{row.cognome} {row.nome}".strip().casefold(),
            f"{row.nome} {row.cognome}".strip().casefold(),
            (row.email or "").casefold(),
        }]
        if len(matches) != 1:
            raise ValueError(f"Consulente non trovato o ambiguo: {clean}")
        return matches[0].id
    return value


def _sheet_rows(worksheet, fields, *, legacy=False):
    header_map = field_by_label(fields)
    if legacy:
        header_map.update({label.casefold(): {"name": name, "label": label, "required": False, "validation": None, "type": "text", "resolver": None} for label, name in LEGACY_HEADERS.items()})
    columns = {}
    for cell in worksheet[1]:
        field = header_map.get(_text(cell.value).casefold())
        if field:
            columns[cell.column] = field
    has_format_row = any("obbligatorio" in _text(cell.value).casefold() or "facoltativo" in _text(cell.value).casefold() for cell in worksheet[2])
    start_row = 3 if has_format_row else 2
    for row_number in range(start_row, worksheet.max_row + 1):
        values = {field["name"]: worksheet.cell(row_number, column).value for column, field in columns.items()}
        if not any(value not in (None, "") for value in values.values()):
            continue
        if _text(values.get("ragione_sociale")) == EXAMPLE_NAME:
            continue
        if row_number == 3 and has_format_row:
            continue
        yield row_number, values, {field["name"]: get_column_letter(column) for column, field in columns.items()}


def parse_workbook(content: bytes, db: Session):
    workbook = load_workbook(BytesIO(content), data_only=True)
    worksheet = workbook["Aziende"] if "Aziende" in workbook.sheetnames else workbook.worksheets[0]
    is_legacy = not any(name in workbook.sheetnames for name in SHEET_SPECS)
    present_sheets = list(workbook.sheetnames)
    if is_legacy and any(_text(cell.value).casefold() == "sedi operative".casefold() for cell in worksheet[1]):
        present_sheets.append("Sedi")
    errors = []
    warnings = []
    if is_legacy:
        warnings.append("Formato legacy rilevato: questo formato è deprecato; scarica il nuovo template multi-foglio con il foglio Sedi.")

    companies = []
    by_piva = {}
    main_fields = importable_company_fields()
    main_by_name = {field["name"]: field for field in main_fields}
    for row_number, raw, columns in _sheet_rows(worksheet, main_fields, legacy=is_legacy):
        normalized = {}
        row_errors = []
        for name, field in main_by_name.items():
            try:
                value = _validate_value(field, raw.get(name))
                if name == "attivo" and value is None:
                    value = True
                if field.get("resolver"):
                    value = _resolve_foreign_key(db, field, value)
                normalized[name] = value
            except (ValueError, TypeError) as exc:
                row_errors.append({"sheet": worksheet.title, "row": row_number, "column": columns.get(name, field["label"]), "message": str(exc)})
        if row_errors:
            errors.extend(row_errors)
            continue
        piva = normalized.get("partita_iva")
        if piva in by_piva:
            errors.append({"sheet": worksheet.title, "row": row_number, "column": columns.get("partita_iva", "Partita IVA"), "message": "Partita IVA duplicata nel file"})
            continue
        normalized.update({"_row": row_number, "sedi_operative": [], "conti_correnti": [], "fund_memberships": []})
        if is_legacy and raw.get("sedi_operative_raw"):
            for chunk in _text(raw["sedi_operative_raw"]).split(";"):
                parts = [part.strip() for part in chunk.split("|")]
                if not any(parts):
                    continue
                parts += [""] * (6 - len(parts))
                normalized["sedi_operative"].append({
                    "nome": parts[0], "tipo": "operativa", "indirizzo": parts[1], "citta": parts[2],
                    "cap": parts[3] or None, "provincia": parts[4].upper() or None,
                    "email": None, "telefono": None, "is_principale": False, "note": parts[5] or None,
                })
        companies.append(normalized)
        by_piva[piva] = normalized

    for sheet_name, fields in SHEET_SPECS.items():
        if sheet_name not in workbook.sheetnames:
            continue
        field_map = {field["name"]: field for field in fields}
        for row_number, raw, columns in _sheet_rows(workbook[sheet_name], fields):
            normalized = {}
            row_errors = []
            for name, field in field_map.items():
                try:
                    normalized[name] = _validate_value(field, raw.get(name))
                except (ValueError, TypeError) as exc:
                    row_errors.append({"sheet": sheet_name, "row": row_number, "column": columns.get(name, field["label"]), "message": str(exc)})
            piva = normalized.pop("partita_iva", None)
            company = by_piva.get(piva)
            if piva and not company:
                row_errors.append({"sheet": sheet_name, "row": row_number, "column": columns.get("partita_iva", "Partita IVA azienda"), "message": "Nessuna azienda valida con questa Partita IVA nel foglio Aziende"})
            if row_errors:
                errors.extend(row_errors)
                continue
            relation = {"Sedi": "sedi_operative", "Conti": "conti_correnti", "Fondi": "fund_memberships"}[sheet_name]
            company[relation].append(normalized)

    valid = []
    for company in companies:
        row_number = company.pop("_row")
        try:
            schemas.AziendaClienteCreate(**company)
        except ValidationError as exc:
            for error in exc.errors():
                field_name = str(error["loc"][0]) if error.get("loc") else "riga"
                field = main_by_name.get(field_name)
                errors.append({"sheet": "Aziende", "row": row_number, "column": field["label"] if field else field_name, "message": error["msg"]})
            continue
        company["_row"] = row_number
        valid.append(company)

    existing_pivas = {
        value for (value,) in db.query(models.AziendaCliente.partita_iva).filter(
            models.AziendaCliente.partita_iva.in_([row["partita_iva"] for row in valid])
        ).all()
    } if valid else set()
    return {
        "format": "legacy" if is_legacy else "multi_sheet",
        "present_sheets": present_sheets,
        "warnings": warnings,
        "errors": errors,
        "rows": valid,
        "summary": {
            "create": sum(1 for row in valid if row["partita_iva"] not in existing_pivas),
            "update": sum(1 for row in valid if row["partita_iva"] in existing_pivas),
            "reject": len({(error["sheet"], error["row"]) for error in errors}),
            "valid": len(valid),
        },
    }


def preview_workbook(content: bytes, db: Session):
    parsed = parse_workbook(content, db)
    return {key: value for key, value in parsed.items() if key not in {"rows", "present_sheets"}}


def import_workbook(content: bytes, db: Session):
    parsed = parse_workbook(content, db)
    report_rows = []
    created = updated = failed = 0
    for row in parsed["rows"]:
        row_number = row.pop("_row")
        piva = row["partita_iva"]
        existing = db.query(models.AziendaCliente).filter(models.AziendaCliente.partita_iva == piva).first()
        try:
            if existing:
                update_row = dict(row)
                for sheet_name, relation in (("Sedi", "sedi_operative"), ("Conti", "conti_correnti"), ("Fondi", "fund_memberships")):
                    if sheet_name not in parsed["present_sheets"]:
                        update_row.pop(relation, None)
                crud.update_azienda_cliente(db, existing.id, schemas.AziendaClienteUpdate(**update_row))
                outcome = "Aggiornata"
                updated += 1
            else:
                crud.create_azienda_cliente(db, schemas.AziendaClienteCreate(**row))
                outcome = "Creata"
                created += 1
            report_rows.append({"row": row_number, "partita_iva": piva, "ragione_sociale": row["ragione_sociale"], "outcome": outcome, "message": "OK"})
        except Exception as exc:
            db.rollback()
            failed += 1
            report_rows.append({"row": row_number, "partita_iva": piva, "ragione_sociale": row.get("ragione_sociale"), "outcome": "Scartata", "message": str(exc)})

    for error in parsed["errors"]:
        report_rows.append({
            "row": error["row"], "partita_iva": "", "ragione_sociale": "",
            "outcome": "Scartata", "message": f"{error['sheet']} / {error['column']}: {error['message']}",
        })
    return {
        "created": created,
        "updated": updated,
        "rejected": parsed["summary"]["reject"] + failed,
        "warnings": parsed["warnings"],
        "errors": parsed["errors"],
        "report_rows": sorted(report_rows, key=lambda item: (item["row"], item["outcome"])),
    }
