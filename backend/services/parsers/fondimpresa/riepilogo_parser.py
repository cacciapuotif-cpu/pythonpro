"""Parser Excel riepilogo piano Fondimpresa."""
import re
from datetime import datetime, date
from typing import Any

from services.parsers.base_parser import BaseDocumentParser


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


def _norm(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", _clean(value).lower()).strip()


def _safe_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _clean(value).replace("€", "").replace(" ", "")
    try:
        return float(text.replace(".", "").replace(",", "."))
    except Exception:
        return None


def _safe_int(value: Any) -> int | None:
    number = _safe_float(value)
    return int(number) if number is not None else None


def _safe_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _clean(value)
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except Exception:
            pass
    return None


def _map_regime(value: Any) -> str | None:
    text = _norm(value)
    if not text:
        return None
    if "de minimis" in text or "deminimis" in text:
        return "de_minimis"
    if "esenzione" in text or "651" in text:
        return "esenzione"
    return _clean(value)


def _header_map(row: tuple[Any, ...]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(row):
        label = _norm(cell)
        if not label:
            continue
        if label in {"n", "num", "numero"}:
            mapping["n"] = idx
        elif "riferimento" in label:
            mapping["riferimento"] = idx
        elif "ragione sociale" in label or label == "azienda":
            mapping["ragione_sociale"] = idx
        elif "codice fiscale" in label:
            mapping["codice_fiscale"] = idx
        elif "partita iva" in label:
            mapping["partita_iva"] = idx
        elif "regime" in label and "aiut" in label:
            mapping["regime_aiuto"] = idx
        elif "dimensione" in label:
            mapping["dimensione"] = idx
        elif label in {"prov", "provincia"} or label.startswith("prov "):
            mapping["provincia"] = idx
        elif "dip" in label:
            mapping["num_dipendenti"] = idx
        elif "finanz" in label and "tot" in label:
            mapping["finanziamento"] = idx
        elif "cofin" in label:
            mapping["cofinanziamento"] = idx
        elif "aziende partecipanti" in label:
            mapping["azienda"] = idx
        elif label == "azioni" or "azione" in label:
            mapping["azione"] = idx
        elif label == "id" or "id azione" in label:
            mapping["id_azione"] = idx
        elif "certificazione" in label:
            mapping["certificazione"] = idx
        elif "tipologia" in label:
            mapping["tipologia"] = idx
        elif label == "ore":
            mapping["ore_totali"] = idx
            mapping["ore"] = idx
        elif "aula" in label:
            mapping["ore_aula"] = idx
        elif "toj" in label or "training on" in label:
            mapping["ore_toj"] = idx
        elif label in {"pax", "partecipanti", "n partecipanti"}:
            mapping["n_partecipanti"] = idx
        elif label == "voce":
            mapping["voce"] = idx
        elif "descrizione" in label:
            mapping["descrizione"] = idx
        elif label == "totale" or label == "tot":
            mapping["totale"] = idx
        elif "nominativo" in label:
            mapping["nominativo"] = idx
        elif label == "ore" and "ore_totali" not in mapping:
            mapping["ore"] = idx
        elif "costo orario" in label:
            mapping["costo_orario"] = idx
        elif "tipo doc" in label:
            mapping["tipo_doc"] = idx
    return mapping


def _get(row: tuple[Any, ...], mapping: dict[str, int], key: str) -> Any:
    idx = mapping.get(key)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


class RiepilogoParser(BaseDocumentParser):
    def parse(self, filepath: str) -> dict:
        warnings: list[str] = []
        try:
            import openpyxl
        except ImportError:
            return _empty(warnings + ["openpyxl non disponibile"])

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as exc:
            return _empty(warnings + [f"Impossibile aprire XLSX: {exc}"])

        result = _empty(warnings)
        result.update(self._parse_metadata(wb))
        result["aziende_beneficiarie"] = self._parse_aziende(wb, warnings)
        result["azioni_formative"] = self._parse_azioni(wb, warnings)
        result["piano_finanziario"] = self._parse_piano_finanziario(wb, warnings)

        result["importo_totale"] = sum(a.get("finanziamento") or 0 for a in result["aziende_beneficiarie"]) or result.get("importo_totale")
        result["contributo_ente"] = result["importo_totale"]
        result["cofinanziamento"] = sum(a.get("cofinanziamento") or 0 for a in result["aziende_beneficiarie"])
        return result

    def _sheet(self, wb, needle: str):
        for name in wb.sheetnames:
            if needle.lower() in name.lower():
                return wb[name]
        return None

    def _parse_metadata(self, wb) -> dict:
        ws = self._sheet(wb, "RIEPILOGO") or wb.active
        meta = {
            "ente": "Fondimpresa",
            "codice_piano": None,
            "titolo_piano": None,
            "soggetto_attuatore": None,
            "importo_totale": None,
            "contributo_ente": None,
            "cofinanziamento": 0.0,
            "data_approvazione": None,
            "determina_numero": None,
            "determina_data": None,
            "cup": None,
            "id_piano_esterno": None,
            "avviso_numero": None,
        }
        label_map = {
            "codice piano": "codice_piano",
            "cup": "cup",
            "id piano": "id_piano_esterno",
            "titolo piano": "titolo_piano",
            "sogg attuatore": "soggetto_attuatore",
            "soggetto attuatore": "soggetto_attuatore",
            "data approvazione": "data_approvazione",
            "data avvio azioni": "data_avvio_azioni",
            "data conclusione": "data_conclusione",
            "avviso": "avviso_numero",
        }
        for row in ws.iter_rows(values_only=True):
            for idx, cell in enumerate(row):
                label = _norm(cell)
                if not label:
                    continue
                for key, target in label_map.items():
                    if key in label:
                        value = None
                        raw_cell = _clean(cell)
                        if ":" in raw_cell:
                            value = raw_cell.split(":", 1)[1].strip()
                        for next_cell in row[idx + 1:]:
                            if value is None and _clean(next_cell):
                                value = next_cell
                                break
                        if value is None:
                            continue
                        if "data" in target:
                            meta[target] = _safe_date(value)
                        else:
                            meta[target] = _clean(value)
        return meta

    def _parse_aziende(self, wb, warnings: list[str]) -> list[dict]:
        ws = self._sheet(wb, "RIEPILOGO")
        if ws is None:
            warnings.append("Sheet RIEPILOGO non trovato")
            return []

        aziende = []
        mapping = None
        for row in ws.iter_rows(values_only=True):
            maybe = _header_map(row)
            if {"n", "ragione_sociale", "codice_fiscale"}.issubset(maybe):
                mapping = maybe
                continue
            if not mapping:
                continue
            n_value = _get(row, mapping, "n")
            if _safe_int(n_value) is None:
                continue
            ragione = _clean(_get(row, mapping, "ragione_sociale"))
            if not ragione:
                continue
            aziende.append({
                "ragione_sociale": ragione,
                "codice_fiscale": _clean(_get(row, mapping, "codice_fiscale")) or None,
                "partita_iva": _clean(_get(row, mapping, "partita_iva")) or None,
                "regime_aiuto": _map_regime(_get(row, mapping, "regime_aiuto")),
                "num_dipendenti": _safe_int(_get(row, mapping, "num_dipendenti")),
                "provincia": (_clean(_get(row, mapping, "provincia")) or None),
                "finanziamento": _safe_float(_get(row, mapping, "finanziamento")) or 0.0,
                "cofinanziamento": _safe_float(_get(row, mapping, "cofinanziamento")) or 0.0,
            })
        if not aziende:
            warnings.append("Nessuna azienda beneficiaria estratta dal RIEPILOGO")
        return aziende

    def _parse_azioni(self, wb, warnings: list[str]) -> list[dict]:
        ws = self._sheet(wb, "PROGETTAZIONE")
        if ws is None:
            warnings.append("Sheet PROGETTAZIONE DETTAGLIO non trovato")
            return []

        azioni = []
        mapping = None
        current_azienda = None
        for row in ws.iter_rows(values_only=True):
            maybe = _header_map(row)
            if "azienda" in maybe and "azione" in maybe:
                mapping = maybe
                continue
            if not mapping:
                continue
            azienda = _clean(_get(row, mapping, "azienda"))
            if azienda:
                current_azienda = azienda
            titolo = _clean(_get(row, mapping, "azione"))
            if not titolo or titolo.lower().startswith("totale"):
                continue
            ore_aula = _safe_int(_get(row, mapping, "ore_aula"))
            ore_toj = _safe_int(_get(row, mapping, "ore_toj"))
            ore_totali = _safe_int(_get(row, mapping, "ore_totali"))
            if ore_totali is None:
                ore_totali = (ore_aula or 0) + (ore_toj or 0) or None
            azioni.append({
                "titolo": titolo,
                "id_azione_esterno": _clean(_get(row, mapping, "id_azione")) or None,
                "azienda_ragione_sociale": current_azienda,
                "ore_totali": ore_totali or 0,
                "ore_aula": ore_aula,
                "ore_toj": ore_toj,
                "n_partecipanti": _safe_int(_get(row, mapping, "n_partecipanti")) or 0,
                "importo": _safe_float(_get(row, mapping, "finanziamento")) or 0.0,
                "certificazione": _norm(_get(row, mapping, "certificazione")) in {"si", "s", "yes", "true"},
            })
        if not azioni:
            warnings.append("Nessuna azione formativa estratta")
        return azioni

    def _parse_piano_finanziario(self, wb, warnings: list[str]) -> list[dict]:
        ws = self._sheet(wb, "PIANO FINANZIARIO")
        if ws is None:
            warnings.append("Sheet PIANO FINANZIARIO non trovato")
            return []

        voci = []
        mapping = None
        current_voce = None
        current_desc = None
        for row in ws.iter_rows(values_only=True):
            maybe = _header_map(row)
            if "voce" in maybe and ("descrizione" in maybe or "totale" in maybe):
                mapping = maybe
                continue
            if not mapping:
                continue
            voce_raw = _clean(_get(row, mapping, "voce"))
            desc_raw = _clean(_get(row, mapping, "descrizione"))
            if voce_raw:
                current_voce = voce_raw
            if desc_raw:
                current_desc = desc_raw
            totale = _safe_float(_get(row, mapping, "totale"))
            ore = _safe_float(_get(row, mapping, "ore"))
            nominativo = _clean(_get(row, mapping, "nominativo")) or None
            costo_orario = _safe_float(_get(row, mapping, "costo_orario"))
            if not current_voce or (totale is None and ore is None and not nominativo):
                continue
            if current_voce.lower().startswith("totale"):
                continue
            voci.append({
                "voce_codice": current_voce,
                "voce_descrizione": current_desc,
                "nominativo": nominativo,
                "ore_previste": ore or 0.0,
                "tariffa_oraria": costo_orario or 0.0,
                "importo_preventivo": totale or 0.0,
                "tipo_doc": _clean(_get(row, mapping, "tipo_doc")) or None,
            })
        if not voci:
            warnings.append("Nessuna voce piano finanziario estratta")
        return voci


def _empty(warnings: list[str]) -> dict:
    return {
        "ente": "Fondimpresa",
        "codice_piano": None,
        "titolo_piano": None,
        "soggetto_attuatore": None,
        "importo_totale": None,
        "contributo_ente": None,
        "cofinanziamento": 0.0,
        "data_approvazione": None,
        "determina_numero": None,
        "determina_data": None,
        "cup": None,
        "id_piano_esterno": None,
        "avviso_numero": None,
        "aziende_beneficiarie": [],
        "azioni_formative": [],
        "piano_finanziario": [],
        "warnings": warnings,
    }
