"""Parser XLSX piano finanziario FAPI — basato sulla struttura reale del documento."""
import re
import logging
import unicodedata
from typing import Any

logger = logging.getLogger(__name__)

# ── mappatura categoria dalla descrizione voce ────────────────────────────────

_CATEGORIA_MAP = [
    ("docenz", "docenza"),
    ("tutor", "tutoraggio"),
    ("progett", "progettazione"),
    ("material", "materiali_didattici"),
    ("aula", "aula"),
    ("attrezzat", "attrezzature"),
    ("certific", "certificazioni"),
    ("viaggio", "viaggi"),
    ("vitto", "viaggi"),
    ("pernott", "viaggi"),
    ("coordinam", "coordinamento"),
    ("direzione", "coordinamento"),
    ("segreteria", "coordinamento"),
    ("amminist", "coordinamento"),
    ("assicuraz", "altro"),
    ("retribuz", "altro"),
    ("oneri", "altro"),
    ("promoz", "altro"),
    ("monitorag", "altro"),
]


def _map_categoria(desc: str | None) -> str:
    if not desc:
        return "altro"
    dl = desc.lower()
    for kw, cat in _CATEGORIA_MAP:
        if kw in dl:
            return cat
    return "altro"


def _safe_float(val) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        v = float(val)
        return None if v == 0 else v
    try:
        s = str(val).strip().replace(",", ".").replace(" ", "")
        v = float(s)
        return None if v == 0 else v
    except Exception:
        return None


def _normalize_match_text(value: str | None) -> str:
    if not value:
        return ""
    normalized = unicodedata.normalize("NFKD", str(value).lower())
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    normalized = re.sub(r"[^a-z0-9]+", " ", normalized)
    return re.sub(r"\s+", " ", normalized).strip()


def _extract_azienda_ordinal(value: str | None) -> int | None:
    if not value:
        return None
    matches = re.findall(r"\b(\d{1,2})\b", str(value))
    if not matches:
        return None
    ordinal = int(matches[-1])
    return ordinal if ordinal > 0 else None


def resolve_modulo_formativo_id(db, project, voce: dict[str, Any]) -> int | None:
    """Collega una voce Excel al modulo formativo FAPI quando il match e' univoco."""
    codice_fapi = getattr(project, "codice_fapi", None)
    if not codice_fapi:
        return None

    ordinal = _extract_azienda_ordinal(voce.get("azienda"))
    if not ordinal:
        return None

    materia_voce = _normalize_match_text(voce.get("materia"))
    if not materia_voce or materia_voce.startswith("tutor"):
        return None

    from models import ModuloFormativo

    codice_progetto_fapi = f"{codice_fapi}{ordinal:02d}"
    candidates = (
        db.query(ModuloFormativo)
        .filter(
            ModuloFormativo.project_id == project.id,
            ModuloFormativo.codice_progetto_fapi == codice_progetto_fapi,
            ModuloFormativo.tipo_attivita == "formativa",
        )
        .all()
    )

    matches = []
    for modulo in candidates:
        materia_modulo = _normalize_match_text(modulo.materia)
        if not materia_modulo:
            continue
        if materia_voce == materia_modulo or materia_voce.startswith(materia_modulo) or materia_modulo.startswith(materia_voce):
            matches.append(modulo)

    return matches[0].id if len(matches) == 1 else None


def _voce_macrovoce(codice: str) -> str:
    c = codice.strip().upper()
    if c.startswith("A"):
        return "A"
    if c.startswith("B"):
        return "B"
    if c.startswith("C"):
        return "C"
    if c.startswith("D"):
        return "D"
    return "A"


def _is_voce_codice(s: str) -> bool:
    """'A.1.1.1', 'B.1.2', 'C.2.2' ecc — lettera + dot + cifre."""
    return bool(re.match(r'^[A-D]\.\d', s.strip()))


def _is_totale_row(b, c) -> bool:
    for v in [b, c]:
        if v and "TOTALE" in str(v).upper():
            return True
    return False


def parse_piano_finanziario(xlsx_path: str) -> dict[str, Any]:
    """
    Parsa XLSX piano finanziario FAPI.
    Struttura sheet 'budget':
    - Col B: VOCE DI SPESA (CATEGORIA X, A.1.1.1, etc.)
    - Col C: sottovoce/descrizione
    - Col D: budget voce
    - Col E: Azienda (sparse — eredita riga sopra)
    - Col F: Materia/Incarico (sparse — eredita riga sopra)
    - Col H: Ore
    - Col J: Totale importo
    """
    warnings: list[str] = []

    try:
        import openpyxl
    except ImportError:
        return {"voci": [], "riepilogo_per_azienda": {}, "totale_ore": 0, "totale_importo": 0, "warnings": ["openpyxl non disponibile"]}

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as exc:
        return {"voci": [], "riepilogo_per_azienda": {}, "totale_ore": 0, "totale_importo": 0, "warnings": [f"Impossibile aprire XLSX: {exc}"]}

    # trova sheet budget
    ws = None
    for name in wb.sheetnames:
        if "budget" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active
        warnings.append(f"Sheet 'budget' non trovato, uso sheet attivo: {ws.title}")

    voci: list[dict] = []
    current_categoria = None
    current_voce_codice = None
    current_voce_desc = None
    current_budget_voce = None
    current_azienda = None
    current_materia = None
    aziende_ordine: list[str] = []  # aziende in ordine di apparizione

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Lettura colonne (0-indexed: B=1, C=2, D=3, E=4, F=5, H=7, J=9)
        col_b = row[1] if len(row) > 1 else None
        col_c = row[2] if len(row) > 2 else None
        col_d = row[3] if len(row) > 3 else None
        col_e = row[4] if len(row) > 4 else None
        col_f = row[5] if len(row) > 5 else None
        col_h = row[7] if len(row) > 7 else None
        col_j = row[9] if len(row) > 9 else None

        b_str = str(col_b).strip() if col_b is not None else ""
        c_str = str(col_c).strip() if col_c is not None else ""

        # Salta righe TOTALE
        if _is_totale_row(b_str, c_str):
            continue

        # CATEGORIA A/B/C/D
        if b_str.startswith("CATEGORIA"):
            current_categoria = b_str
            current_azienda = None
            current_materia = None
            continue

        # Voce di spesa (A.1.1.1, B.1.2, etc.)
        if b_str and _is_voce_codice(b_str):
            current_voce_codice = b_str
            current_voce_desc = c_str or None
            current_budget_voce = _safe_float(col_d)
            current_azienda = None   # reset per nuova voce
            current_materia = None
            # se la stessa riga ha ore, non saltiamo
            # (cade nel blocco ore sotto)

        # Aggiorna azienda (col E non vuota)
        e_str = str(col_e).strip() if col_e is not None else ""
        if e_str:
            current_azienda = e_str
            if e_str not in aziende_ordine:
                aziende_ordine.append(e_str)

        # Aggiorna materia (col F non vuota)
        f_str = str(col_f).replace("\n", " ").strip() if col_f is not None else ""
        if f_str:
            current_materia = f_str

        # Override azienda per righe "tutor X N": mappa N → azienda N-esima
        if f_str and not e_str:
            m = re.search(r'(\d+)\s*$', f_str)
            if m and f_str.lower().startswith("tutor"):
                idx = int(m.group(1)) - 1
                if 0 <= idx < len(aziende_ordine):
                    current_azienda = aziende_ordine[idx]

        # Registra voce se ha ore e voce_codice
        ore = _safe_float(col_h)
        if ore and current_voce_codice:
            importo = _safe_float(col_j)
            voci.append({
                "categoria": current_categoria,
                "voce_codice": current_voce_codice,
                "voce_descrizione": current_voce_desc,
                "budget_voce": current_budget_voce,
                "azienda": current_azienda,
                "materia": current_materia,
                "ore_previste": ore,
                "importo_totale": importo,
            })

    # ── Riepilogo per azienda ─────────────────────────────────────────────────
    riepilogo: dict[str, dict] = {}
    totale_ore = 0.0
    totale_importo = 0.0

    for v in voci:
        az = v["azienda"] or "Generale"
        if az not in riepilogo:
            riepilogo[az] = {"ore_totali": 0.0, "importo_totale": 0.0, "voci": []}
        riepilogo[az]["ore_totali"] += v["ore_previste"] or 0
        riepilogo[az]["importo_totale"] += v["importo_totale"] or 0
        riepilogo[az]["voci"].append(v)
        totale_ore += v["ore_previste"] or 0
        totale_importo += v["importo_totale"] or 0

    if not voci:
        warnings.append("Nessuna voce trovata nello sheet budget")

    return {
        "voci": voci,
        "riepilogo_per_azienda": riepilogo,
        "totale_ore": round(totale_ore, 2),
        "totale_importo": round(totale_importo, 2),
        "warnings": warnings,
    }
